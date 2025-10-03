import pandas as pd
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def create_coinbase_forecast():
    """Create the one-tab Q3 2025 Coinbase revenue forecast Excel file."""
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Q3-2025 Coinbase Forecast"
    
    # Define colors and styles
    input_fill = PatternFill(start_color="2222FF", end_color="2222FF", fill_type="solid")
    calc_fill = PatternFill(start_color="0000FF", end_color="0000FF", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                   top=Side(style='thin'), bottom=Side(style='thin'))
    
    # Scenario dropdown (A2)
    ws['A1'] = "Q3 2025 Coinbase Revenue Forecast"
    ws['A1'].font = header_font
    ws['A1'].fill = PatternFill(start_color="0000FF", end_color="0000FF", fill_type="solid")
    
    ws['A2'] = "Scenario"
    ws['B2'] = "Base"
    
    # Data validation cell
    ws['A3'] = "Data Validation: Base, Bull, Bear"
    
    # === INPUTS SECTION (Columns A-E) ===
    ws['A4'] = "INPUTS"
    ws['A4'].font = header_font
    ws['A4'].fill = input_fill
    
    # Transaction Inputs
    ws['A5'] = "Mgmt July Txn Revenue"
    ws['B5'] = "$360,000,000"
    
    ws['A6'] = "Aug Base Coinbase Notional ($B)"
    ws['B6'] = "72"
    
    ws['A7'] = "Sep Base Coinbase Notional ($B)"
    ws['B7'] = "80"
    
    ws['A8'] = "Blended Take Rate"
    ws['B8'] = "0.0025"
    
    ws['A9'] = "Sentiment Jul"
    ws['B9'] = "1.00"
    
    ws['A10'] = "Sentiment Aug"
    ws['B10'] = "1.018"
    
    ws['A11'] = "Sentiment Sep"
    ws['B11'] = "1.015"
    
    # S&S Inputs
    ws['A12'] = ""
    ws['A13'] = "S&S INPUTS"
    ws['A13'].font = header_font
    ws['A13'].fill = input_fill
    
    ws['B12'] = ""
    ws['B13'] = "Interest"
    ws['C13'] = "Staking"
    ws['D13'] = "Custody"
    
    ws['A14'] = "Balance/Units"
    ws['B14'] = "25000000000"
    ws['C14'] = "33000000"
    ws['D14'] = "150000000000"
    
    ws['A15'] = "Rate/Price"
    ws['B15'] = "0.0540"
    ws['C15'] = "3500"
    ws['D15'] = "0.00013"
    
    ws['A16'] = "Share/Take"
    ws['B16'] = "0.22"
    ws['C16'] = "0.115"
    ws['D16'] = ""
    
    ws['A17'] = "Other"
    ws['D17'] = "-1500000"
    
    # === CALCULATIONS SECTION (Columns F-K) ===
    ws['F4'] = "CALCULATIONS"
    ws['F4'].font = header_font
    ws['F4'].fill = calc_fill
    
    ws['F5'] = "Transactions"
    ws['G5'] = ""
    ws['H5'] = "S&S Components"
    
    ws['F6'] = "Txn_Jul"
    ws['G6'] = "=B5"
    
    ws['F7'] = "Txn_Aug"
    ws['G7'] = "=B6*1000000000*B8*B10"
    
    ws['F8'] = "Txn_Sep"
    ws['G8'] = "=B7*1000000000*B8*B11"
    
    ws['F9'] = "Txn_Total"
    ws['G9'] = "=SUM(G6:G8)"
    
    ws['H6'] = "Interest"
    ws['I6'] = "=B14*B15*B16+28000000000*B15*0.0035"
    
    ws['H7'] = "Staking"
    ws['I7'] = "=(C14*C15*0.040/4)*C16"
    
    ws['H8'] = "Custody"
    ws['I8'] = "=D14*D15"
    
    ws['H9'] = "Other"
    ws['I9'] = "=D17"
    
    ws['H10'] = "S&S_Total"
    ws['I10'] = "=SUM(I6:I9)"
    
    ws['G11'] = ""
    ws['I11'] = ""
    
    ws['F12'] = "Total Revenue"
    ws['J12'] = "=G9+I10"
    
    # === SCENARIOS SECTION ===
    ws['L4'] = "SCENARIOS"
    ws['L4'].font = header_font
    ws['L4'].fill = PatternFill(start_color="FF9900", end_color="FF9900", fill_type="solid")
    
    ws['L5'] = "Base"
    ws['M5'] = "Bull"
    ws['N5'] = "Bear"
    
    ws['L6'] = "Aug Notional +0%"
    ws['M6'] = "Aug Notional +10%"
    ws['N6'] = "Aug Notional -10%"
    
    ws['L7'] = "Sep Notional +0%"
    ws['M7'] = "Sep Notional +10%"
    ws['N7'] = "Sep Notional -10%"
    
    ws['L8'] = "Take Rate 0.0025"
    ws['M8'] = "Take Rate 0.0027"
    ws['N8'] = "Take Rate 0.0023"
    
    ws['L9'] = "Sentiment factor"
    ws['M9'] = "Sentiment +0.02"
    ws['N9'] = "Sentiment -0.02"
    
    ws['L10'] = "Base Rev"
    ws['M10'] = "Bull Rev"
    ws['N10'] = "Bear Rev"
    
    ws['L11'] = "=J12"
    ws['M11'] = "=J12*1.08"  # Approximate bull case
    ws['N11'] = "=J12*0.92"  # Approximate bear case
    
    # === SENSITIVITY TABLE ===
    ws['F15'] = "SENSITIVITY ANALYSIS"
    ws['F15'].font = header_font
    ws['F15'].fill = PatternFill(start_color="009900", end_color="009900", fill_type="solid")
    
    ws['F16'] = "Take Rate"
    ws['G16'] = "-5bps"
    ws['H16'] = "Base"
    ws['I16'] = "+5bps"
    
    ws['F17'] = "Delta Revenue"
    ws['G17'] = "=-((B8-0.0025)*2000000000000)"
    ws['H17'] = "=J12"
    ws['I17'] = "=(B8+0.0025-0.0025)*2000000000000"
    
    ws['F18'] = ""
    ws['F19'] = "Notional ±10%"
    ws['G19'] = "Revenue Δ"
    ws['F20'] = "Aug"
    ws['G20'] = "=J12*0.038"
    ws['F21'] = "Sep"
    ws['G21'] = "=J12*0.062"
    
    # Apply borders to all cells
    for row in range(1, 25):
        for col in range(1, 15):
            cell = ws.cell(row=row, column=col)
            cell.border = border
            
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = min(max_length + 2, 20)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save the file
    wb.save("model/coinbase_forecast.xlsx")
    print("Created coinbase_forecast.xlsx successfully!")
    
    return "model/coinbase_forecast.xlsx"

if __name__ == "__main__":
    create_coinbase_forecast()
